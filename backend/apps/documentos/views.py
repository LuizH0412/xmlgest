from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import Documento, ItemDocumento
from core.permissions import PaginacaoPadrao
from .serializers import DocumentoSerializer
from core.permissions import IsAdmin
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from apps.empresas.models import Empresa
from .services import extrair_dados
from .filters import DocumentoFilter
from django.utils import timezone
import pandas as pd
import re
from apps.empresas.authentication import ColetorUser
import os
from django.http import FileResponse, HttpResponse
from brazilfiscalreport.danfe import Danfe
import io
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from decimal import Decimal
import zipfile
from apps.documentos.models import ExportacaoXml
import zipfile
from io import BytesIO
from apps.documentos.gerar_xml_service import gerar_xml_nota
from apps.auditoria import service as auditoria


class DocumentoViewSet(viewsets.ModelViewSet):
    lookup_field = 'chave_acesso'
    pagination_class = PaginacaoPadrao
    filterset_class = DocumentoFilter
    queryset = Documento.objects.all()
    serializer_class = DocumentoSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.action == 'download_export':
            return []
        if self.action in ['list', 'retrieve', 'create', 'upload', 'enviar_xmls']:
            return [IsAuthenticated()]
        return [IsAdmin()]

    def destroy(self, request, *args, **kwargs):
        if request.user.perfil != 'admin':
            return Response({'detail': 'Sem permissão.'}, status=status.HTTP_403_FORBIDDEN)
        
        documento = self.get_object()
        if documento.caminho_arquivo:
            caminho = os.path.join('media', documento.caminho_arquivo)
            if os.path.exists(caminho):
                os.remove(caminho)

        documento.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    

    @action(detail=False, methods=['post'], url_path='validar-upload', parser_classes=[MultiPartParser])
    def validar_upload(self, request):
        arquivo = request.FILES.get('arquivo')
        empresa_id = request.query_params.get('empresa') 

        if not arquivo:
            return Response({'detail': 'Nenhum arquivo enviado.'}, status=400)

        if not empresa_id:
            return Response({'detail': 'Parâmetro empresa é obrigatório.'}, status=400)

        try:
            empresa_atual = Empresa.objects.get(id=empresa_id)
        except Empresa.DoesNotExist:
            return Response({'detail': 'Empresa não encontrada.'}, status=404)

        try:
            dados = extrair_dados(arquivo)
        except ValueError as e:
            return Response({'detail': str(e)}, status=400)

        # Valida se o CNPJ do XML pertence à empresa que está sendo acessada
        cnpj_xml = dados['cnpj_emitente']
        if cnpj_xml != empresa_atual.cnpj:
            return Response({
                'status': 'cnpj_invalido',
                'cnpj': cnpj_xml,
                'label': 'CNPJ não pertence a esta empresa'
            }, status=200)

        existente = Documento.objects.filter(chave_acesso=dados['chave_acesso']).first()

        if not existente:
            return Response({
                'status': 'novo',
                'label': 'Não existe na base',
                'dados': dados,  
            })

        divergencias = []
        if str(existente.valor_total) != str(dados['valor_total']):
            divergencias.append(f"Valor: {existente.valor_total} → {dados['valor_total']}")
        if str(existente.numero_nota) != str(dados['numero_nota']):
            divergencias.append(f"Número: {existente.numero_nota} → {dados['numero_nota']}")
        if existente.serie != dados['serie']:
            divergencias.append(f"Série: {existente.serie} → {dados['serie']}")

        if divergencias:
            return Response({
                'status': 'divergente',
                'label': 'Existente com valores divergentes',
                'divergencias': divergencias,
                'dados': dados,  
            })

        return Response({
            'status': 'existente',
            'label': 'Já existe na base',
            'dados': dados,  
        })

    @action(detail=False, methods=['post'], url_path='upload', parser_classes=[MultiPartParser])
    def upload(self, request):
        arquivo = request.FILES.get('arquivo')

        if not arquivo:
            return Response(
                {'detail': 'Nenhum arquivo enviado.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            dados = extrair_dados(arquivo)
        except ValueError as e:
            return Response(
                {'detail': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verifica duplicata
        if Documento.objects.filter(chave_acesso=dados['chave_acesso']).exists():
            return Response(
                {'detail': 'Documento já existe no sistema.'},
                status=status.HTTP_409_CONFLICT
            )

        # Identifica a empresa pelo CNPJ
        try:
            empresa = Empresa.objects.get(cnpj=dados['cnpj_emitente'])
        except Empresa.DoesNotExist:
            return Response(
                {'detail': 'Empresa não encontrada para o CNPJ informado.'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Salva o arquivo
        data = dados['data_emissao']
        caminho = f"xmls/{empresa.codigo_interno}/{dados['tipo']}/{data.year}/{data.month:02d}/{arquivo.name}"
        caminho_completo = os.path.join('media', caminho)
        os.makedirs(os.path.dirname(caminho_completo), exist_ok=True)
        with open(caminho_completo, 'wb+') as destino:
            for chunk in arquivo.chunks():
                destino.write(chunk)

        # Salva no banco
        enviado_por = None if isinstance(request.user, ColetorUser) else request.user

        documento = Documento.objects.create(
            empresa=empresa,
            chave_acesso=dados['chave_acesso'],
            tipo=dados['tipo'],
            numero_nota=dados['numero_nota'],
            serie=dados['serie'],
            data_emissao=dados['data_emissao'],
            valor_total=dados['valor_total'],
            status=dados['status'],
            caminho_arquivo=caminho,
            enviado_por=enviado_por,
        )
        auditoria.registrar(
            request,
            acao='upload_documento',
            detalhes=f'Chave: {documento.chave_acesso}',
            empresa=empresa,
        )

        ItemDocumento.objects.bulk_create([
            ItemDocumento(documento=documento, empresa=empresa, **item)
            for item in dados.get('itens', [])
        ])
        

        return Response(
            {'detail': 'Documento enviado com sucesso.', 'id': documento.id},
            status=status.HTTP_201_CREATED
        )
    
    @action(detail=False, methods=['post'], url_path='preview', parser_classes=[MultiPartParser])
    def preview(self, request):
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            return Response({'detail': 'Nenhum arquivo enviado.'}, status=400)
        try:
            dados = extrair_dados(arquivo)
            return Response(dados)
        except ValueError as e:
            return Response({'detail': str(e)}, status=400)

    @action(detail=True, methods=['get'], url_path='download-xml')
    def download_xml(self, request, chave_acesso=None):
        try:
            documento = Documento.objects.get(chave_acesso=chave_acesso)
        except Documento.DoesNotExist:
            return Response({'detail': 'Documento não encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        caminho_completo = os.path.join('media', documento.caminho_arquivo)

        if not os.path.exists(caminho_completo):
            return Response(
                {'detail': 'Arquivo XML não encontrado.'},
                status=status.HTTP_404_NOT_FOUND
            )

        return FileResponse(
            open(caminho_completo, 'rb'),
            as_attachment=True,
            filename=f"{documento.chave_acesso}.xml",
            content_type='application/xml'
        )

    @action(detail=True, methods=['get'], url_path='download-pdf')
    def download_pdf(self, request, chave_acesso=None):
        try:
            documento = Documento.objects.get(chave_acesso=chave_acesso)
        except Documento.DoesNotExist:
            return Response({'detail': 'Documento não encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        caminho_completo = os.path.join('media', documento.caminho_arquivo)

        if not os.path.exists(caminho_completo):
            return Response({'detail': 'Arquivo XML não encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        with open(caminho_completo, 'rb') as f:
            xml_content = f.read()

        pdf_buffer = io.BytesIO()
        danfe = Danfe(xml=xml_content)
        danfe.output(pdf_buffer)
        pdf_buffer.seek(0)

        return HttpResponse(
            pdf_buffer,
            content_type='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename="{chave_acesso}.pdf"'
            }
        )

    @action(detail=False, methods=['get'], url_path='relatorio/excel')
    def relatorio_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Documentos"

        ws.merge_cells('A1:G1')
        ws['A1'] = 'Relatório de Documentos Fiscais'
        ws['A1'].font = Font(bold=True, size=14, color='1a1a1a')
        ws['A1'].fill = PatternFill('solid', fgColor='FACC15')
        ws['A1'].alignment = Alignment(horizontal='center')

        empresa_id = request.query_params.get('empresa', '')
        tipo = request.query_params.get('tipo', '')
        data_inicio = request.query_params.get('data_emissao__gte', '')
        data_fim = request.query_params.get('data_emissao__lte', '')
        filtros = f"Empresa ID: {empresa_id or 'Todas'} | Tipo: {tipo or 'Todos'} | Período: {data_inicio or '...'} até {data_fim or '...'}"
        ws['A2'] = filtros
        ws['A2'].font = Font(size=9, color='666666')
        ws.merge_cells('A2:G2')

        ws.append([])

        headers = ['Número', 'Tipo', 'Série', 'Data Emissão', 'Valor Total', 'Status', 'Chave de Acesso']
        ws.append(headers)
        header_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F2937')
            cell.alignment = Alignment(horizontal='center')

        total_valor = Decimal('0')
        total_autorizados = 0
        total_cancelados = 0

        for doc in queryset:
            valor = doc.valor_total or Decimal('0')
            total_valor += valor
            if doc.status == 'autorizado':
                total_autorizados += 1
            elif doc.status == 'cancelado':
                total_cancelados += 1

            ws.append([
                doc.numero_nota,
                doc.tipo,
                doc.serie,
                doc.data_emissao.strftime('%d/%m/%Y'),
                float(valor),
                doc.status,
                doc.chave_acesso,
            ])

        ws.append([])

        totais = [
            ('Total de Documentos', queryset.count()),
            ('Autorizados', total_autorizados),
            ('Cancelados', total_cancelados),
            ('Valor Total', float(total_valor)),
        ]
        for label, valor in totais:
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            cell = ws.cell(row=row, column=2, value=valor)
            if label == 'Valor Total':
                cell.number_format = 'R$ #,##0.00'
                cell.font = Font(bold=True, color='15803D')

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 50

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="relatorio_documentos.xlsx"'}
        )

    @action(detail=False, methods=['get'], url_path='relatorio/pdf')
    def relatorio_pdf(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        buffer = io.BytesIO()
        pdf_doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)

        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('title', fontSize=16, fontName='Helvetica-Bold',
                                    textColor=colors.HexColor('#1F2937'), spaceAfter=4)
        elements.append(Paragraph('Relatório de Documentos Fiscais', title_style))

        empresa_id = request.query_params.get('empresa', '')
        tipo = request.query_params.get('tipo', '')
        data_inicio = request.query_params.get('data_emissao__gte', '')
        data_fim = request.query_params.get('data_emissao__lte', '')
        filtros_txt = f"Empresa ID: {empresa_id or 'Todas'} | Tipo: {tipo or 'Todos'} | Período: {data_inicio or '...'} até {data_fim or '...'}"
        sub_style = ParagraphStyle('sub', fontSize=8, textColor=colors.HexColor('#6B7280'), spaceAfter=12)
        elements.append(Paragraph(filtros_txt, sub_style))

        header = ['Número', 'Tipo', 'Série', 'Emissão', 'Valor Total', 'Status']
        rows = [header]

        total_valor = Decimal('0')
        total_autorizados = 0
        total_cancelados = 0

        for doc in queryset:
            valor = doc.valor_total or Decimal('0')
            total_valor += valor
            if doc.status == 'autorizado':
                total_autorizados += 1
            elif doc.status == 'cancelado':
                total_cancelados += 1

            rows.append([
                doc.numero_nota,
                doc.tipo,
                doc.serie,
                doc.data_emissao.strftime('%d/%m/%Y'),
                f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                doc.status,
            ])

        col_widths = [3*cm, 2.5*cm, 2*cm, 3*cm, 4*cm, 3*cm]
        table = Table(rows, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F9FAFB'), colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#E5E7EB')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))

        total_rows = [
            ['Total de Documentos', str(queryset.count())],
            ['Autorizados', str(total_autorizados)],
            ['Cancelados', str(total_cancelados)],
            ['Valor Total', f"R$ {total_valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ]
        total_table = Table(total_rows, colWidths=[6*cm, 4*cm])
        total_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.HexColor('#FACC15')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('TEXTCOLOR', (1, -1), (1, -1), colors.HexColor('#15803D')),
            ('FONTNAME', (1, -1), (1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(total_table)

        pdf_doc.build(elements)
        buffer.seek(0)

        return HttpResponse(
            buffer,
            content_type='application/pdf',
            headers={'Content-Disposition': 'attachment; filename="relatorio_documentos.pdf"'}
        )

    @action(detail=False, methods=['post'], url_path='inconsistencias', parser_classes=[MultiPartParser])
    def inconsistencias(self, request):
        arquivo = request.FILES.get('arquivo')
        empresa_id = request.query_params.get('empresa')

        if not arquivo:
            return Response({'detail': 'Nenhum arquivo enviado.'}, status=400)
        if not empresa_id:
            return Response({'detail': 'Parâmetro empresa é obrigatório.'}, status=400)

        try:
            empresa = Empresa.objects.get(id=empresa_id)
        except Empresa.DoesNotExist:
            return Response({'detail': 'Empresa não encontrada.'}, status=404)

        try:
            conteudo = arquivo.read()
            df_raw = pd.read_excel(BytesIO(conteudo), engine='xlrd', header=None)

            header_row = None
            for i, row in df_raw.iterrows():
                row_str = ' '.join([str(v).upper() for v in row.values if pd.notna(v)])
                if 'NUMERO NOTA FISCAL' in row_str:
                    header_row = i
                    break

            if header_row is None:
                return Response({'detail': 'Formato de planilha não reconhecido. Verifique se é uma planilha da SEFAZ.'}, status=400)

            cnpj_planilha = None
            for i, row in df_raw.iterrows():
                if i >= header_row:
                    break
                for val in row.values:
                    val_str = str(val)
                    if 'CNPJ' in val_str.upper():
                        match = re.search(r'\d{2}[\.\-]?\d{3}[\.\-]?\d{3}[\/]?\d{4}[\-]?\d{2}', val_str)
                        if match:
                            cnpj_planilha = ''.join(filter(str.isdigit, match.group()))
                            break
                if cnpj_planilha:
                    break

            if cnpj_planilha:
                cnpj_empresa = ''.join(filter(str.isdigit, empresa.cnpj))
                if cnpj_planilha != cnpj_empresa:
                    return Response({
                        'detail': f'A planilha pertence ao CNPJ {cnpj_planilha}, mas a empresa selecionada possui o CNPJ {cnpj_empresa}. Importe a planilha correta.'
                    }, status=400)

            df = pd.read_excel(BytesIO(conteudo), engine='xlrd', header=header_row)
            df.columns = [str(c).strip().upper() for c in df.columns]

            col_map = {}
            for col in df.columns:
                if 'NUMERO' in col and 'NOTA' in col:
                    col_map[col] = 'numero'
                elif col.startswith('SÉR') or col.startswith('SER'):
                    col_map[col] = 'serie'
                elif 'SITUA' in col:
                    col_map[col] = 'situacao'
                elif 'CHAVE' in col:
                    col_map[col] = 'chave_acesso'
                elif 'DATA' in col and 'EMISS' in col:
                    col_map[col] = 'data_emissao'
                elif ('VALR' in col or 'VALOR' in col) and 'NOTA' in col:
                    col_map[col] = 'valor_total'
                elif 'PROTOCOLO' in col:
                    col_map[col] = 'protocolo'
                elif 'NATUREZA' in col:
                    col_map[col] = 'natureza_operacao'

            df = df.rename(columns=col_map)
            colunas_necessarias = ['numero', 'serie', 'situacao']
            faltando_cols = [c for c in colunas_necessarias if c not in df.columns]
            if faltando_cols:
                return Response({'detail': f'Colunas não encontradas na planilha: {faltando_cols}. Verifique o formato.'}, status=400)

            cols_disponiveis = [v for v in col_map.values() if v in df.columns]
            df = df[cols_disponiveis].dropna(subset=['numero', 'serie'])
            df['numero'] = pd.to_numeric(df['numero'], errors='coerce')
            df = df.dropna(subset=['numero'])
            df['numero'] = df['numero'].astype(int)
            df['serie'] = pd.to_numeric(df['serie'], errors='coerce').fillna(1).astype(int).astype(str)
            df['situacao'] = df['situacao'].astype(str).str.strip()

        except Exception as e:
            return Response({'detail': f'Erro ao ler planilha: {str(e)}'}, status=400)

        data_min = None
        data_max = None
        if 'data_emissao' in df.columns:
            datas = pd.to_datetime(df['data_emissao'], dayfirst=True, errors='coerce').dropna()
            if not datas.empty:
                data_min = datas.min().date()
                data_max = datas.max().date()

        docs_qs = Documento.objects.filter(empresa=empresa)
        if data_min and data_max:
            docs_qs = docs_qs.filter(data_emissao__gte=data_min, data_emissao__lte=data_max)

        docs_sistema = docs_qs.values(
            'numero_nota', 'serie', 'chave_acesso', 'status', 'data_emissao', 'valor_total'
        )
        sistema_map = {
            (str(d['serie']), str(int(d['numero_nota']))): d
            for d in docs_sistema
        }

        sefaz_notas = set()
        faltando_no_sistema = []
        for _, row in df.iterrows():
            key = (str(row['serie']), str(row['numero']))
            sefaz_notas.add(key)
            if key not in sistema_map:
                faltando_no_sistema.append({
                    'numero': str(row['numero']),
                    'serie': str(row['serie']),
                    'situacao': row['situacao'],
                    'chave_acesso': str(row.get('chave_acesso', '')),
                    'data_emissao': str(row.get('data_emissao', '')),
                    'valor_total': str(row.get('valor_total', '')),
                    'protocolo': str(row.get('protocolo', '')),           
                    'natureza_operacao': str(row.get('natureza_operacao', '')),  
                })

        extras_no_sistema = []
        for (serie, numero), doc in sistema_map.items():
            if (serie, numero) not in sefaz_notas:
                extras_no_sistema.append({
                    'numero': numero,
                    'serie': serie,
                    'status': doc['status'],
                    'chave_acesso': doc['chave_acesso'],
                    'data_emissao': str(doc['data_emissao']),
                    'valor_total': str(doc['valor_total']),
                })

        # ── Gaps na sequência da SEFAZ ────────────────────────────────────────
        gaps = []
        series_sefaz = {}
        for (serie, numero) in sefaz_notas:
            series_sefaz.setdefault(serie, []).append(int(numero))

        for serie, numeros in series_sefaz.items():
            numeros_sorted = sorted(numeros)
            for i in range(len(numeros_sorted) - 1):
                atual = numeros_sorted[i]
                proximo = numeros_sorted[i + 1]
                if proximo - atual > 1:
                    for faltante in range(atual + 1, proximo):
                        gaps.append({
                            'serie': serie,
                            'numero': faltante,
                            'entre': f'{atual} e {proximo}',
                        })

        # ── Gaps na sequência do SISTEMA ─────────────────────────────────────
        gaps_sistema = []
        series_sistema = {}
        for (serie, numero) in sistema_map.keys():
            series_sistema.setdefault(serie, []).append(int(numero))

        for serie, numeros in series_sistema.items():
            numeros_sorted = sorted(numeros)
            for i in range(len(numeros_sorted) - 1):
                atual = numeros_sorted[i]
                proximo = numeros_sorted[i + 1]
                if proximo - atual > 1:
                    for faltante in range(atual + 1, proximo):
                        gaps_sistema.append({
                            'serie': serie,
                            'numero': faltante,
                            'entre': f'{atual} e {proximo}',
                        })

        # ── Valor total da planilha ───────────────────────────────────────────
        valor_total_sefaz = Decimal('0')
        if 'valor_total' in df.columns:
            total = pd.to_numeric(
                df['valor_total'].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False),
                errors='coerce'
            ).sum()
            valor_total_sefaz = Decimal(str(round(total, 2)))

        return Response({
            'empresa': empresa.nome_fantasia,
            'total_sefaz': len(df),
            'valor_total_sefaz': str(valor_total_sefaz),
            'periodo': {
                'inicio': str(data_min) if data_min else None,
                'fim': str(data_max) if data_max else None,
            },
            'faltando_no_sistema': faltando_no_sistema,
            'extras_no_sistema': extras_no_sistema,
            'gaps_sequencia': gaps,
            'gaps_sistema': gaps_sistema,
        })

    @action(detail=False, methods=['post'], url_path='enviar-xmls')
    def enviar_xmls(self, request):
        empresa_id = request.query_params.get('empresa')
        if not empresa_id:
            return Response({'detail': 'Parâmetro empresa é obrigatório.'}, status=400)

        try:
            empresa = Empresa.objects.get(id=empresa_id)
        except Empresa.DoesNotExist:
            return Response({'detail': 'Empresa não encontrada.'}, status=404)

        if not empresa.email_contabilidade:
            return Response({'detail': 'Empresa não possui email de contabilidade cadastrado.'}, status=400)

        data_inicio = request.query_params.get('data_emissao__gte')
        data_fim = request.query_params.get('data_emissao__lte')
        tipo = request.query_params.get('tipo')
        serie = request.query_params.get('serie')
        numero_nota = request.query_params.get('numero_nota')

        from apps.documentos.tasks import enviar_xmls_empresa
        enviar_xmls_empresa.delay(
            empresa_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            tipo=tipo,
            serie=serie,
            numero_nota=numero_nota,
        )

        return Response({'detail': 'Envio iniciado com sucesso.'})

    @action(detail=False, methods=['get'], url_path='download-xmls')
    def download_xmls(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        zip_buffer = io.BytesIO()
        arquivos_incluidos = 0

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for doc in queryset:
                caminho_completo = os.path.join('media', doc.caminho_arquivo)
                if os.path.exists(caminho_completo):
                    zf.write(caminho_completo, arcname=os.path.basename(caminho_completo))
                    arquivos_incluidos += 1

        if arquivos_incluidos == 0:
            return Response({'detail': 'Nenhum arquivo encontrado.'}, status=404)

        zip_buffer.seek(0)

        return HttpResponse(
            zip_buffer,
            content_type='application/zip',
            headers={'Content-Disposition': 'attachment; filename="xmls.zip"'}
        )

    @action(detail=False, methods=['get'], url_path='download-export/(?P<token>[^/.]+)', permission_classes=[])
    def download_export(self, request, token=None):
        try:
            exportacao = ExportacaoXml.objects.get(token=token)
        except ExportacaoXml.DoesNotExist:
            return Response({'detail': 'Link inválido.'}, status=404)

        if not os.path.exists(exportacao.caminho_arquivo):
            return Response({'detail': 'Arquivo não encontrado.'}, status=404)

        return FileResponse(
            open(exportacao.caminho_arquivo, 'rb'),
            as_attachment=True,
            filename=os.path.basename(exportacao.caminho_arquivo),
            content_type='application/zip'
        )
    

    @action(detail=False, methods=['post'], url_path='gerar-xml')
    def gerar_xml(self, request):
        """
        Gera XMLs reconstituídos para notas ausentes no sistema.
 
        Retorna:
        - XML direto (application/xml) se apenas uma nota sem erros
        - ZIP (application/zip) se múltiplas notas
        """
        from apps.documentos.gerar_xml_service import gerar_xml_nota
        from decimal import Decimal
 
        empresa_id = request.data.get('empresa_id')
        notas = request.data.get('notas', [])
 
        if not empresa_id or not notas:
            return Response(
                {'detail': 'empresa_id e notas são obrigatórios.'},
                status=status.HTTP_400_BAD_REQUEST
            )
 
        try:
            empresa = Empresa.objects.get(id=empresa_id)
        except Empresa.DoesNotExist:
            return Response({'detail': 'Empresa não encontrada.'}, status=status.HTTP_404_NOT_FOUND)
 
        resultados = []
        erros = []
 
        for nota in notas:
            numero_nota = nota.get('numero_nota')
            try:
                xml_bytes, chave = gerar_xml_nota(
                    empresa=empresa,
                    numero_nota=numero_nota,
                    serie=nota.get('serie'),
                    valor_total_sefaz=nota.get('valor_total'),
                    data_emissao=nota.get('data_emissao'),
                    chave_acesso=nota.get('chave_acesso'),
                    protocolo=nota.get('protocolo'),
                    natureza_operacao=nota.get('natureza_operacao'),
                )
 
                if Documento.objects.filter(chave_acesso=chave).exists():
                    erros.append({'nota': numero_nota, 'erro': 'Documento já existe no sistema.'})
                    continue
 
                # Salva arquivo no disco
                from datetime import datetime as dt
                data = dt.strptime(nota.get('data_emissao'), '%Y-%m-%d').date()
                nome_arquivo = f"{chave}-reconstituido.xml"
                caminho = f"xmls/{empresa.codigo_interno}/NFe/{data.year}/{data.month:02d}/{nome_arquivo}"
                caminho_completo = os.path.join('media', caminho)
                os.makedirs(os.path.dirname(caminho_completo), exist_ok=True)
 
                with open(caminho_completo, 'wb') as f:
                    f.write(xml_bytes)
 
                # Salva Documento no banco
                enviado_por = None if isinstance(request.user, ColetorUser) else request.user
                documento = Documento.objects.create(
                    empresa=empresa,
                    chave_acesso=chave,
                    tipo='NFe',
                    numero_nota=numero_nota,
                    serie=nota.get('serie'),
                    data_emissao=data,
                    valor_total=Decimal(str(nota.get('valor_total'))),
                    status='autorizado',
                    caminho_arquivo=caminho,
                    enviado_por=enviado_por,
                )
 
                # Salva itens via extrair_dados do XML gerado
                from apps.documentos.services import extrair_dados
                arquivo_virtual = BytesIO(xml_bytes)
                arquivo_virtual.name = nome_arquivo
                dados = extrair_dados(arquivo_virtual)
                ItemDocumento.objects.bulk_create([
                    ItemDocumento(documento=documento, empresa=empresa, **item)
                    for item in dados.get('itens', [])
                ])
 
                resultados.append({
                    'nota': numero_nota,
                    'chave': chave,
                    'xml': xml_bytes,
                    'nome': nome_arquivo,
                })
 
            except Exception as e:
                import traceback
                traceback.print_exc()
                erros.append({'nota': numero_nota, 'erro': str(e)})
 
        if not resultados:
            return Response(
                {'detail': 'Nenhum XML gerado.', 'erros': erros},
                status=status.HTTP_400_BAD_REQUEST
            )
 
        # Retorno: XML direto (1 nota) ou ZIP (múltiplas)
        if len(resultados) == 1:
            r = resultados[0]
            from django.http import HttpResponse as HR
            response = HR(r['xml'], content_type='application/xml')
            response['Content-Disposition'] = f'attachment; filename="{r["nome"]}"'
            return response
 
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for r in resultados:
                zf.writestr(r['nome'], r['xml'])
        zip_buffer.seek(0)
 
        from django.http import HttpResponse as HR
        response = HR(zip_buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="xmls_reconstituidos_{empresa.codigo_interno}.zip"'
        return response